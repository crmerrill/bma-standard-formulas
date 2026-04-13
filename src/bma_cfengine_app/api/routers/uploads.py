from __future__ import annotations

from typing import Optional

from fastapi import APIRouter, File, Form, HTTPException, Query, UploadFile
from fastapi.encoders import jsonable_encoder
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from ...orchestrator.mapping import (
    apply_mapping,
    auto_infer_mappings,
    profile_dataframe,
    sanitize_field_mappings,
)
from ...orchestrator.rates import rates_preflight, save_rates_file
from ...orchestrator.run_service import compute_tape_stats
from ...orchestrator.strats import available_strat_dimensions, compute_strat, summarize_tape, summarize_unique_values
from ...orchestrator.dq_normalizer import (
    DqMapping,
    detect_dq_pattern,
    materialize_dq_columns,
    suggest_dq_mapping,
)
from ...orchestrator.tape_repair import (
    apply_repair,
    available_repairs,
    diagnose_tape,
    preview_repair,
)
from ...storage import run_store
from ..models import FieldMapping, RatesPreflightResponse, TapeProfile, TapeStats, UploadResponse

router = APIRouter(tags=["uploads"])


def _load_mapped_df(upload_id: str, mapping_id: str | None = None):
    """Load the upload DataFrame, applying mapping if a mapping_id is provided."""
    df, file_name = run_store.load_upload_df(upload_id)
    if mapping_id:
        try:
            mapping_data = run_store.load_mapping(upload_id, mapping_id)
            mappings = sanitize_field_mappings(
                [FieldMapping(**m) for m in mapping_data["mappings"]]
            )
            df = apply_mapping(df, mappings)
        except FileNotFoundError:
            pass
    return df, file_name


@router.post("/uploads", response_model=UploadResponse)
async def upload_tape(file: UploadFile = File(...), display_name: str | None = Form(None)):
    upload_id = run_store.new_upload_id()
    content = await file.read()
    file_name = file.filename or "tape.csv"
    run_store.save_upload(upload_id, file_name, content, display_name=display_name)

    df, _ = run_store.load_upload_df(upload_id)
    stored_meta = run_store.load_upload_metadata(upload_id)
    return UploadResponse(
        upload_id=upload_id,
        file_name=file_name,
        display_name=str(stored_meta.get("display_name") or "").strip() or file_name,
        row_count=len(df),
        column_count=len(df.columns),
    )


@router.get("/uploads")
async def list_uploads():
    return {"items": run_store.list_uploads()}


class UploadRenameRequest(BaseModel):
    display_name: str


@router.patch("/uploads/{upload_id}")
async def rename_upload(upload_id: str, body: UploadRenameRequest):
    try:
        meta = run_store.set_upload_display_name(upload_id, body.display_name)
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=f"Upload {upload_id!r} not found") from e
    return {
        "upload_id": upload_id,
        "display_name": str(meta.get("display_name") or "").strip(),
    }


@router.get("/uploads/{upload_id}/mappings")
async def list_upload_mappings(upload_id: str):
    return {"upload_id": upload_id, "items": run_store.list_mappings(upload_id)}


@router.get("/uploads/{upload_id}/profile", response_model=TapeProfile)
async def get_profile(upload_id: str):
    df, file_name = run_store.load_upload_df(upload_id)
    file_size = run_store.raw_upload_byte_size(upload_id)
    return profile_dataframe(df, upload_id, file_name, file_size)


@router.get("/uploads/{upload_id}/auto-map", response_model=list[FieldMapping])
async def auto_map(upload_id: str):
    df, _ = run_store.load_upload_df(upload_id)
    return sanitize_field_mappings(auto_infer_mappings(list(df.columns)))


@router.get("/uploads/{upload_id}/stats", response_model=TapeStats)
async def tape_stats(upload_id: str, mapping_id: Optional[str] = Query(None)):
    df, _ = _load_mapped_df(upload_id, mapping_id)
    return compute_tape_stats(df)


@router.get("/uploads/{upload_id}/preview")
async def tape_preview(
    upload_id: str,
    mapping_id: Optional[str] = Query(None),
    limit: int = 100,
):
    df, _ = _load_mapped_df(upload_id, mapping_id)
    preview = df.head(limit)
    return {
        "columns": list(preview.columns),
        "rows": preview.fillna("").to_dict("records"),
        "total_rows": len(df),
        "showing": len(preview),
    }


@router.get("/uploads/{upload_id}/tape-summary")
async def tape_summary(upload_id: str, mapping_id: Optional[str] = Query(None)):
    import math
    df, _ = _load_mapped_df(upload_id, mapping_id)
    summary_df = summarize_tape(df)
    rows = summary_df.to_dict("records")
    for row in rows:
        for k, v in row.items():
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                row[k] = None
    payload = {
        "columns": list(summary_df.columns),
        "rows": rows,
        "row_count": len(rows),
    }
    return jsonable_encoder(payload)


@router.get("/uploads/{upload_id}/unique-values")
async def unique_values(upload_id: str, mapping_id: Optional[str] = Query(None)):
    import math
    df, _ = _load_mapped_df(upload_id, mapping_id)
    uv_df = summarize_unique_values(df)
    rows = uv_df.to_dict("records")
    for row in rows:
        for k, v in row.items():
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                row[k] = None
    payload = {
        "columns": list(uv_df.columns),
        "rows": rows,
        "row_count": len(rows),
    }
    return jsonable_encoder(payload)


@router.get("/uploads/{upload_id}/strat-dimensions")
async def strat_dimensions(upload_id: str, mapping_id: Optional[str] = Query(None)):
    df, _ = _load_mapped_df(upload_id, mapping_id)
    return available_strat_dimensions(df)


class StratRequest(BaseModel):
    group_by: str | list[str]
    mapping_id: str | None = None
    max_buckets: int = 10
    filter: dict[str, str] | None = None


@router.post("/uploads/{upload_id}/strats")
async def compute_strats(upload_id: str, req: StratRequest):
    df, _ = _load_mapped_df(upload_id, req.mapping_id)
    result = compute_strat(
        df, req.group_by,
        max_buckets=req.max_buckets,
        filter_=req.filter,
    )
    return {
        "group_by": req.group_by,
        "columns": list(result.columns),
        "rows": result.to_dict("records"),
        "row_count": len(result),
    }


class ExportStratsRequest(BaseModel):
    dimensions: list[str]
    mapping_id: str | None = None
    max_buckets: int = 10
    format: str = "xlsx"


@router.post("/uploads/{upload_id}/strats-export")
async def export_strats(upload_id: str, req: ExportStratsRequest):
    import io
    df, _ = _load_mapped_df(upload_id, req.mapping_id)
    tables = {dim: compute_strat(df, dim, max_buckets=req.max_buckets) for dim in req.dimensions}

    if req.format == "csv":
        buf = io.StringIO()
        for dim, tbl in tables.items():
            buf.write(f"# Stratification: {dim}\n")
            tbl.to_csv(buf, index=False)
            buf.write("\n")
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode()),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=strats_{upload_id[:12]}.csv"},
        )

    buf = io.BytesIO()
    with __import__("openpyxl").Workbook() as _:
        pass
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for dim, tbl in tables.items():
        safe_name = dim[:31].replace("/", "_").replace("\\", "_")
        ws = wb.create_sheet(title=safe_name)
        ws.append(list(tbl.columns))
        for _, row in tbl.iterrows():
            ws.append([v for v in row])
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=strats_{upload_id[:12]}.xlsx"},
    )


# ---------------------------------------------------------------------------
# Data quality / repair
# ---------------------------------------------------------------------------


@router.get("/uploads/{upload_id}/diagnose")
async def diagnose(upload_id: str, mapping_id: Optional[str] = Query(None)):
    df, _ = _load_mapped_df(upload_id, mapping_id)
    report = diagnose_tape(df)
    repairs = available_repairs(df)
    return {**report, "available_repairs": repairs}


@router.get("/uploads/{upload_id}/repair-preview")
async def repair_preview_endpoint(
    upload_id: str,
    rule_id: str,
    mapping_id: Optional[str] = Query(None),
    limit: int = 20,
):
    df, _ = _load_mapped_df(upload_id, mapping_id)
    return preview_repair(df, rule_id, limit)


class ApplyRepairRequest(BaseModel):
    rule_id: str
    mapping_id: str | None = None


@router.post("/uploads/{upload_id}/apply-repair")
async def apply_repair_endpoint(upload_id: str, req: ApplyRepairRequest):
    df, _ = _load_mapped_df(upload_id, req.mapping_id)
    df_fixed, count = apply_repair(df, req.rule_id)
    run_store.save_working_copy(upload_id, df_fixed)

    return {
        "rule_id": req.rule_id,
        "rows_fixed": count,
        "has_working_copy": True,
        "message": f"Applied '{req.rule_id}': {count} value(s) computed and saved to working copy.",
    }


@router.post("/uploads/{upload_id}/revert")
async def revert_to_raw(upload_id: str):
    run_store.revert_to_raw(upload_id)
    return {"message": "Reverted to original raw file.", "has_working_copy": False}


@router.get("/uploads/{upload_id}/status")
async def upload_status(upload_id: str):
    return {
        "upload_id": upload_id,
        "has_working_copy": run_store.has_working_copy(upload_id),
    }


# ---------------------------------------------------------------------------
# DQ normalization
# ---------------------------------------------------------------------------


@router.get("/uploads/{upload_id}/dq-detect")
async def dq_detect(upload_id: str, mapping_id: Optional[str] = Query(None)):
    """Auto-detect the DQ pattern in the tape and return a mapping suggestion."""
    df, _ = _load_mapped_df(upload_id, mapping_id)
    mapping = suggest_dq_mapping(df)
    return mapping.model_dump()


class DqApplyRequest(BaseModel):
    mapping_id: str | None = None
    dq_mapping: DqMapping


@router.post("/uploads/{upload_id}/dq-apply")
async def dq_apply(upload_id: str, req: DqApplyRequest):
    """Apply a DQ mapping to the tape, materializing canonical DQ columns."""
    df, _ = _load_mapped_df(upload_id, req.mapping_id)
    enriched = materialize_dq_columns(df, req.dq_mapping)
    run_store.save_working_copy(upload_id, enriched)

    # Persist the DQ mapping alongside the column mapping
    dq_path = run_store.upload_dir(upload_id) / "dq_mapping.json"
    dq_path.write_text(req.dq_mapping.model_dump_json(indent=2))

    canonical_cols = ["dlq_status", "days_past_due", "is_fc", "is_reo"]
    found = [c for c in canonical_cols if c in enriched.columns]

    return {
        "upload_id": upload_id,
        "pattern": req.dq_mapping.pattern,
        "columns_added": found,
        "row_count": len(enriched),
        "has_working_copy": True,
        "message": f"DQ normalization applied ({req.dq_mapping.pattern}): "
                   f"{len(found)} canonical column(s) added.",
    }


# ---------------------------------------------------------------------------
# Run preflight (tape readiness check)
# ---------------------------------------------------------------------------

REQUIRED_FOR_RUN = [
    "loan_id", "origination_date", "asof_date", "original_balance",
    "current_balance", "rate_margin", "original_term", "remaining_term",
]


@router.get("/uploads/{upload_id}/run-preflight")
async def run_preflight(upload_id: str, mapping_id: Optional[str] = Query(None)):
    df, _ = _load_mapped_df(upload_id, mapping_id)
    blocking: list[str] = []
    warnings: list[str] = []

    for field in REQUIRED_FOR_RUN:
        if field not in df.columns:
            blocking.append(f"Required field '{field}' is not mapped")
            continue
        nan_count = int(df[field].isna().sum())
        if nan_count > 0:
            blocking.append(
                f"'{field}' has {nan_count} missing value(s) — fix on the Tape View page before running"
            )

    return {
        "ready": len(blocking) == 0,
        "blocking": blocking,
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Rates file
# ---------------------------------------------------------------------------


@router.post("/uploads/{upload_id}/rates")
async def upload_rates(upload_id: str, file: UploadFile = File(...)):
    content = await file.read()
    file_name = file.filename or "rates.csv"
    save_rates_file(upload_id, file_name, content)
    return {"upload_id": upload_id, "file_name": file_name}


@router.get("/uploads/{upload_id}/rates-preflight", response_model=RatesPreflightResponse)
async def rates_preflight_endpoint(
    upload_id: str,
    mapping_id: Optional[str] = Query(None),
):
    return rates_preflight(upload_id, mapping_id)
